import streamlit as st
st.set_page_config(layout="wide")
import pandas as pd
from openpyxl import load_workbook

st.title("REGISTER DATA ENTRY")
file_path = "data.xlsx"


def display_excel(file_path):
    df = pd.read_excel(file_path)
    st.dataframe(df)

def delete_rows_ws():
    wb = load_workbook("data.xlsx")
    ws = wb.active
    row_max = ws.max_row
    if row_max <=6:
        pass
        st.error("NO MORE ROWS TO DELETE")
    else:
        ws.delete_rows(row_max,1)
        wb.save(file_path)
        st.success("DELETED THE ROW")
        display_excel(file_path)

with st.form("data_entry_form"):

    # Balono = st.number_input("**BALOPASANA KRAMANK**", min_value=1,max_value=200)
    Name = st.text_input("**NAME**")
    Standard = st.number_input("**STANDARD**", min_value=1,max_value=10)
    Gender = st.selectbox("**GENDER**", ["M", "F"])
    Phone = st.text_input("**PHONE**")
    Shaka = st.selectbox("SHAKHA", ["SAI MANDIR","JESHTH NAGARIK MAANCH","GANESH MAIDAN","PARMARTH NIKETAN","SANKALP SIDDHI","BEST COLONY","DATTA MANDIR","MANGALMURTI GANESH MANDIR","PANDUMASTER"])
    Address = st.text_input("**ADDRESS**")

    submitted = st.form_submit_button("**SUBMIT**")   
st.markdown(" ")

col1,col2 = st.columns([5,5])

if col1.button("Delete Row"):
    delete_rows_ws()

if submitted:

    workbook = load_workbook(file_path)
    sheet = workbook.active
    row = sheet.max_row

    if Name=="" or Phone=="" or Address=="":
        st.error("NO NAME OR PHONE NUMBER OR ADDRESS")
    else:
        next_row = sheet.max_row + 1
        serial_number = sheet.max_row - 5
        sheet[f"A{next_row}"] = serial_number
        sheet[f"B{next_row}"] = serial_number
        sheet[f"C{next_row}"] = Name
        sheet[f"D{next_row}"] = Standard
        sheet[f"E{next_row}"] = Gender
        sheet[f"F{next_row}"] = Phone
        sheet[f"G{next_row}"] = Shaka
        sheet[f"H{next_row}"] = Address
        workbook.save(file_path)
        
        st.success(f"{Name} {Standard} {Gender} {Phone} {Shaka} {Address} Added Sucessfully")
        display_excel(file_path)
else:
    display_excel(file_path)

st.markdown(" ")

import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from openpyxl import load_workbook
from io import BytesIO

# JavaScript to toggle cell value between "P" and None on click, with restrictions
cell_click_js = JsCode("""
function(e) {
    let cell = e.api.getFocusedCell();
    let rowIndex = cell.rowIndex;
    let colId = cell.column.colId;
    let colIndex = cell.column.instanceId;
    // Define the columns A to H as 0 to 7 (zero-indexed)
    if (rowIndex >4 && colIndex >= 8) {
        let rowNode = e.api.getRowNode(rowIndex);
        let cellValue = rowNode.data[colId];
        if (cellValue === "P") {
            rowNode.setDataValue(colId, "A");
        } else {
            rowNode.setDataValue(colId, "P");
        }
    }
}
""")
with st.container(border=True):
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
st.markdown(" ")

if uploaded_file:
    
    # Load the uploaded Excel file into a DataFrame
    content = uploaded_file.read()
    df = pd.read_excel(BytesIO(content), engine='openpyxl')
    
    # st.write("Original DataFrame:")
    # st.dataframe(df)

    # Set up AgGrid options
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(editable=True)
    gb.configure_grid_options(onCellClicked=cell_click_js)
    grid_options = gb.build()

    # Display the DataFrame with AgGrid
    # st.write("Click on any cell to toggle its value between 'P' and empty (except rows 1-6 and columns A-H)")
    grid_response = AgGrid(
        df,
        gridOptions=grid_options,
        editable=True,
        height=500,
        allow_unsafe_jscode=True,  # Set it to True to allow custom JsCode
        reload_data=True
    )

    # Updated DataFrame after clicking and editing
    updated_df = pd.DataFrame(grid_response['data'])

    # st.write("Updated DataFrame:")
    # st.dataframe(updated_df)

    if st.button('SAVE CHANGES'):
        # Load the workbook "data.xlsx"
        wb = load_workbook('data.xlsx')
        sheet = wb.active
        
        # Update the Excel sheet with updated DataFrame values
        for r_idx, row in updated_df.iterrows():
            for c_idx, value in enumerate(row):
                # Convert r_idx and c_idx to integers if they are coming as strings
                rr = int(r_idx) + 2
                cc = int(c_idx) + 1
                sheet.cell(row=rr, column=cc, value=value)
        
        # Save the changes to "data.xlsx"
        wb.save('data.xlsx')
        
        st.success("SUCESSFULLY SAVED")
