import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode
from openpyxl import load_workbook
from io import BytesIO

# JavaScript to toggle cell value between "P" and None on click
cell_click_js = JsCode("""
function(e) {
    let cell = e.api.getFocusedCell();
    let rowIndex = cell.rowIndex;
    let colId = cell.column.colId;
    let rowNode = e.api.getRowNode(rowIndex);
    let cellValue = rowNode.data[colId];
    if (cellValue === "P") {
        rowNode.setDataValue(colId, "A");
    } else {
        rowNode.setDataValue(colId, "P");
    }
}
""")

st.title("Excel Cell Click and Modify")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    # Load the uploaded Excel file into a DataFrame
    content = uploaded_file.read()
    df = pd.read_excel(BytesIO(content), engine='openpyxl')
    
    st.write("Original DataFrame:")
    st.dataframe(df)

    # Set up AgGrid options
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(editable=True)
    gb.configure_grid_options(onCellClicked=cell_click_js)
    grid_options = gb.build()

    # Display the DataFrame with AgGrid
    st.write("Click on any cell to toggle its value between 'P' and empty")
    grid_response = AgGrid(
        df,
        gridOptions=grid_options,
        editable=True,
        height=500,
        allow_unsafe_jscode=True,  # Set it to True to allow custom JsCode
        reload_data=True
    )

    # Updated DataFrame after clicking and editing
    updated_df = pd.DataFrame(grid_response['data'])

    st.write("Updated DataFrame:")
    st.dataframe(updated_df)

    if st.button('Save Changes to data.xlsx'):
        # Load the workbook "data.xlsx"
        wb = load_workbook('data.xlsx')
        sheet = wb.active
        
        # Update the Excel sheet with updated DataFrame values
        for r_idx, row in updated_df.iterrows():
            for c_idx, value in enumerate(row):
                # Convert r_idx and c_idx to integers if they are coming as strings
                rr = int(r_idx) + 2
                cc = int(c_idx) + 1
                sheet.cell(row=rr, column=cc, value=value)
        
        # Save the changes to "data.xlsx"
        wb.save('data.xlsx')
        
        st.success("Changes saved to data.xlsx.")
